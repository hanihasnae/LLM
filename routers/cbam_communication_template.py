"""
routers/cbam_communication_template.py

Official EU CBAM Communication Template Excel Generator
Conforme : Règlement UE 2023/956 + Règlement d'exécution UE 2025/2620

⚠️  MANDATORY: Tab name "Summary_Communication" must NOT be renamed.
    Renaming causes file rejection by the EU CBAM Portal.
"""

from fastapi import APIRouter, HTTPException, Query, Depends
from fastapi.responses import StreamingResponse
from dependencies import get_current_user
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import get_connection

router = APIRouter(prefix="/cbam", tags=["CBAM Communication Template"])

# ── EU Official Color Palette ──────────────────────────────────────────────────
EU_BLUE       = "003399"   # EU flag blue
EU_YELLOW     = "FFCC00"   # EU flag yellow
HEADER_DARK   = "1F3864"   # Dark navy for main headers
MANDATORY_BG  = "FFF2CC"   # Light yellow — cells the operator must fill
READONLY_BG   = "F2F2F2"   # Light gray — auto-calculated cells (read-only)
GREEN_OK      = "E2EFDA"   # Light green — compliant status
WHITE         = "FFFFFF"

# ── Helpers DB ─────────────────────────────────────────────────────────────────

def _get_profile(conn, user_id: int) -> dict:
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM company_profile WHERE user_id = %s ORDER BY id LIMIT 1",
        (user_id,)
    )
    row = cur.fetchone()
    cur.close()
    if not row:
        raise HTTPException(
            status_code=404,
            detail="Profil entreprise non configuré — allez dans Paramètres"
        )
    return dict(row)


def _get_emissions_summary(conn, year: int) -> dict:
    cur = conn.cursor()
    cur.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN e.scope = 1 THEN e.co2_kg ELSE 0 END), 0) AS scope1_kg,
            COALESCE(SUM(CASE WHEN e.scope = 2 THEN e.co2_kg ELSE 0 END), 0) AS scope2_kg,
            COALESCE(SUM(
                CASE WHEN a.source = 'electricity' THEN a.quantity ELSE 0 END
            ), 0) AS elec_kwh,
            COUNT(DISTINCT a.id) AS activity_count
        FROM emissions e
        JOIN activities a ON e.activity_id = a.id
        WHERE a.actif = true
          AND e.actif = true
          AND EXTRACT(YEAR FROM a.date) = %s
    """, (year,))
    row = cur.fetchone()
    cur.close()
    if not row:
        return {"scope1_kg": 0.0, "scope2_kg": 0.0, "elec_kwh": 0.0, "activity_count": 0}
    return {
        "scope1_kg":      float(row["scope1_kg"]),
        "scope2_kg":      float(row["scope2_kg"]),
        "elec_kwh":       float(row["elec_kwh"]),
        "activity_count": int(row["activity_count"]),
    }

# ── Style Builders ─────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _font(bold=False, color=None, size=10, italic=False) -> Font:
    return Font(bold=bold, color=color or "000000", size=size, italic=italic)


def _border_thin() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _styled(cell, value, fill_hex=None, bold=False, color=None,
            size=10, h_align="left", wrap=False, border=True, italic=False):
    cell.value     = value
    cell.font      = _font(bold=bold, color=color, size=size, italic=italic)
    cell.alignment = _align(h=h_align, wrap=wrap)
    if fill_hex:
        cell.fill  = _fill(fill_hex)
    if border:
        cell.border = _border_thin()
    return cell

# ── Public Style Helpers ──────────────────────────────────────────────────────

def thin_border() -> Border:
    """Thin gray border (#BFBFBF) for regular cells."""
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def medium_border() -> Border:
    """Medium navy border (#1F3864) for section containers."""
    s = Side(style="medium", color="1F3864")
    return Border(left=s, right=s, top=s, bottom=s)


def header_cell(cell, text, bg_color, fg_color, bold=True, size=10, wrap=False, halign="left"):
    """Header styling: Arial font, solid fill, centered vertically, thin border."""
    cell.value     = text
    cell.font      = Font(name="Arial", bold=bold, color=fg_color, size=size)
    cell.fill      = _fill(bg_color)
    cell.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    cell.border    = thin_border()


def input_cell(cell, value, bg_color=MANDATORY_BG, halign="left", fmt=None):
    """
    Operator input cell: blue Arial font on light yellow background.
    Pass fmt="#,##0.00" for numeric formatting.
    """
    cell.value     = value
    cell.font      = Font(name="Arial", size=10, color="003399")
    cell.fill      = _fill(bg_color)
    cell.alignment = Alignment(horizontal=halign, vertical="center")
    cell.border    = thin_border()
    if fmt:
        cell.number_format = fmt


def calculated_cell(cell, formula_or_value, halign="center", fmt=None, bg=READONLY_BG):
    """
    Read-only / calculated cell: black Arial on gray background.
    If formula_or_value starts with '=', it is kept as a formula string.
    """
    cell.value     = formula_or_value
    cell.font      = Font(name="Arial", size=10, color="000000")
    cell.fill      = _fill(bg)
    cell.alignment = Alignment(horizontal=halign, vertical="center")
    cell.border    = thin_border()
    if fmt:
        cell.number_format = fmt


def label_cell(cell, text, bold=False, color="000000", bg=WHITE, halign="left", size=10):
    """Descriptive label: plain Arial, configurable color/size/bg."""
    cell.value     = text
    cell.font      = Font(name="Arial", bold=bold, color=color, size=size)
    cell.fill      = _fill(bg)
    cell.alignment = Alignment(horizontal=halign, vertical="center")
    cell.border    = thin_border()


def merge_and_style_section_title(ws, row, start_col, end_col, text, bg_color):
    """
    Merge start_col→end_col on the given row, apply bold white header styling.
    Row height is set to 28.
    """
    start = get_column_letter(start_col)
    end   = get_column_letter(end_col)
    ws.merge_cells(f"{start}{row}:{end}{row}")
    header_cell(ws.cell(row, start_col), text,
                bg_color=bg_color, fg_color=WHITE,
                bold=True, size=11, halign="left")
    ws.row_dimensions[row].height = 28


def note_cell(ws, row, start_col, end_col, text, bg="FFF9C4", fg_color="7B5800"):
    """
    Merge start_col→end_col, render italic note text on light yellow background.
    Row height is set to 16.
    """
    start = get_column_letter(start_col)
    end   = get_column_letter(end_col)
    ws.merge_cells(f"{start}{row}:{end}{row}")
    cell           = ws.cell(row, start_col)
    cell.value     = text
    cell.font      = Font(name="Arial", italic=True, size=9, color=fg_color)
    cell.fill      = _fill(bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border    = thin_border()
    ws.row_dimensions[row].height = 16


# ── Section Builders ──────────────────────────────────────────────────────────

def build_section_1_installation_details(ws, start_row: int, profile: dict) -> int:
    """
    SECTION 1 — INSTALLATION DETAILS
    Identifies the operator and installation (IR 2025/2547, Annex IV §1.1).

    Args:
        ws         : openpyxl Worksheet
        start_row  : first row to write into
        profile    : dict from _get_profile(conn)  — company_profile row

    Returns:
        next_row   : start_row + 10 (title + header + 8 data rows)

    Column layout:
        A(5)  — Ref      B(32) — Field       C(22) — Value
        D–E(22 each, merged) — Source        F(24) — Required?
    """
    # ── Column widths (set once, idempotent if called again) ───────────────
    for col_letter, width in [("A", 5), ("B", 32), ("C", 40),
                               ("D", 22), ("E", 22), ("F", 24)]:
        ws.column_dimensions[col_letter].width = width

    # ── Row 0 offset : Section title ──────────────────────────────────────
    merge_and_style_section_title(
        ws, start_row, 1, 6,
        "SECTION 1 — INSTALLATION DETAILS",
        bg_color=HEADER_DARK,
    )

    # ── Row +1 : Column headers ────────────────────────────────────────────
    h_row = start_row + 1
    for col, text in enumerate(["#", "Field", "Value", "Source", "", "Required?"], 1):
        header_cell(ws.cell(h_row, col), text,
                    bg_color="2E4057", fg_color=WHITE,
                    bold=True, size=9, halign="center")
    # Merge D+E header
    ws.merge_cells(f"D{h_row}:E{h_row}")
    ws.row_dimensions[h_row].height = 28

    # ── Data rows (8 fields) ───────────────────────────────────────────────
    # Each entry: (ref, field_label, value, source_note, required)
    # Fields marked "← TO ADD IN DB" don't exist in company_profile yet.
    rows_data = [
        (
            "1.1",
            "Operator / Company Name",
            profile.get("nom_entreprise") or "",
            "company_profile.nom_entreprise",
            "MANDATORY",
        ),
        (
            "1.2",
            "Full Address (English)",
            profile.get("adresse_en") or "[company_profile.adresse_en] ← TO ADD IN DB",
            "company_profile.adresse_en",
            "MANDATORY",
        ),
        (
            "1.3",
            "Installation Name",
            profile.get("installation_name") or "[company_profile.installation_name] ← TO ADD IN DB",
            "company_profile.installation_name",
            "MANDATORY",
        ),
        (
            "1.4",
            "GPS Coordinates (lat, lon)",
            (
                f"{profile['gps_latitude']}, {profile['gps_longitude']}"
                if profile.get("gps_latitude") and profile.get("gps_longitude")
                else "[company_profile.gps_latitude, gps_longitude] ← TO ADD IN DB"
            ),
            "company_profile.gps_latitude / gps_longitude",
            "MANDATORY",
        ),
        (
            "1.5",
            "UN/LOCODE",
            "MACAS",                               # Fixed — Morocco / Casablanca
            "Fixed value (Morocco/Casablanca)",
            "MANDATORY",
        ),
        (
            "1.6",
            "CBAM Registry ID",
            profile.get("cbam_registry_id") or "[company_profile.cbam_registry_id] ← TO ADD IN DB",
            "company_profile.cbam_registry_id",
            "Optional",
        ),
        (
            "1.7",
            "Registration Number (RC / SIRET)",
            profile.get("registration_number") or "[company_profile.registration_number] ← TO ADD IN DB",
            "company_profile.registration_number",
            "MANDATORY",
        ),
        (
            "1.8",
            "Country of Installation",
            "Morocco",                             # Fixed
            "Fixed value",
            "MANDATORY",
        ),
    ]

    for i, (ref, field, value, source, required) in enumerate(rows_data):
        r = start_row + 2 + i
        ws.row_dimensions[r].height = 34

        # Col A — Ref
        label_cell(ws.cell(r, 1), ref,
                   bold=True, color=HEADER_DARK, bg="EEF2FF", halign="center")

        # Col B — Field label
        label_cell(ws.cell(r, 2), field,
                   bold=True, color="000000", bg="F8FAFF")

        # Col C — Value (input cell — operator fills or already from DB)
        is_placeholder = "← TO ADD IN DB" in str(value)
        c_cell = ws.cell(r, 3)
        c_cell.value     = value
        c_cell.font      = Font(name="Arial", size=10, color="003399",
                                italic=is_placeholder)
        c_cell.fill      = _fill(MANDATORY_BG)
        c_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c_cell.border    = thin_border()

        # Col D–E — Source note (merged, gray, small)
        ws.merge_cells(f"D{r}:E{r}")
        src_cell = ws.cell(r, 4)
        src_cell.value     = source
        src_cell.font      = Font(name="Arial", size=9, color="595959", italic=True)
        src_cell.fill      = _fill(WHITE)
        src_cell.alignment = Alignment(horizontal="left", vertical="center")
        src_cell.border    = thin_border()

        # Col F — Required status
        if required == "MANDATORY":
            req_bg, req_fg = MANDATORY_BG, "E65100"   # orange
        else:
            req_bg, req_fg = GREEN_OK, "276221"        # green

        f_cell = ws.cell(r, 6)
        f_cell.value     = required
        f_cell.font      = Font(name="Arial", size=9, bold=True, color=req_fg)
        f_cell.fill      = _fill(req_bg)
        f_cell.alignment = Alignment(horizontal="center", vertical="center")
        f_cell.border    = thin_border()

    # ── Footer note ────────────────────────────────────────────────────────
    note_row = start_row + 10
    note_cell(
        ws, note_row, 1, 6,
        "ℹ️  Fields marked '← TO ADD IN DB' require adding the column to company_profile "
        "table. Fixed values (UN/LOCODE, Country) are pre-filled per Moroccan installations.",
    )

    return note_row + 1   # next available row for the next section


def build_section_2_summary_of_products(ws, start_row: int, profile: dict, raw: dict) -> int:
    """
    SECTION 2 — SUMMARY OF PRODUCTS
    Lists all CN codes exported with specific emissions and EU benchmarks.

    Args:
        ws         : openpyxl Worksheet
        start_row  : first row to write into
        profile    : dict from _get_profile(conn)
        raw        : dict from _get_emissions_summary(conn, year)

    Returns:
        next_row   : start_row + 6 (title + header + 3 product rows + note)

    Columns:
        A(4)  Ref    B(14) CN Code    C(28) Description
        D(18) Production(t)    E(18) Direct(tCO₂e/t)    F(18) Indirect(tCO₂e/t)
        G(18) Total (formula)  H(16) EU Benchmark        I(20) Compliance (formula)
    """
    HEADER_MID  = "2E5496"
    ORANGE_WARN = "FCE4D6"

    # Column widths
    for col_letter, width in [("A", 4), ("B", 14), ("C", 28),
                               ("D", 18), ("E", 18), ("F", 18),
                               ("G", 18), ("H", 16), ("I", 20)]:
        ws.column_dimensions[col_letter].width = width

    # ── Section title ──────────────────────────────────────────────────────
    merge_and_style_section_title(
        ws, start_row, 1, 9,
        "SECTION 2 — SUMMARY OF PRODUCTS",
        bg_color=HEADER_DARK,
    )

    # ── Column headers ─────────────────────────────────────────────────────
    h_row = start_row + 1
    headers = [
        "#",
        "CN Code",
        "Product Description (EN)",
        "Annual Production\n(tonnes)",
        "Direct Specific\nEmissions\n(tCO₂e/t) — Scope 1",
        "Indirect Specific\nEmissions\n(tCO₂e/t) — Scope 2",
        "Total Specific\nEmissions\n(tCO₂e/t)",
        "EU Benchmark\n(tCO₂e/t)",
        "Compliance\nStatus",
    ]
    for col, text in enumerate(headers, 1):
        header_cell(ws.cell(h_row, col), text,
                    bg_color=HEADER_MID, fg_color=WHITE,
                    bold=True, size=9, halign="center", wrap=True)
    ws.row_dimensions[h_row].height = 42

    # ── Derive real values for row 1 from DB ───────────────────────────────
    production  = float(profile.get("production_annuelle_tonnes") or 0)
    scope1_t    = raw.get("scope1_kg", 0) / 1000.0
    scope2_t    = raw.get("scope2_kg", 0) / 1000.0
    scope1_spec = round(scope1_t / production, 4) if production > 0 else None
    scope2_spec = round(scope2_t / production, 4) if production > 0 else None
    cn_code     = profile.get("cn_code") or ""
    secteur     = profile.get("secteur") or ""
    has_data    = raw.get("activity_count", 0) > 0

    # 3 product placeholder rows (row 1 = real data if available, 2-3 = blank)
    product_rows = [
        {
            "ref":        "2.1",
            "cn_code":    cn_code,
            "desc":       secteur,
            "production": production if production > 0 else None,
            "direct":     scope1_spec if has_data else None,
            "indirect":   scope2_spec if has_data else None,
        },
        {"ref": "2.2", "cn_code": None, "desc": None, "production": None, "direct": None, "indirect": None},
        {"ref": "2.3", "cn_code": None, "desc": None, "production": None, "direct": None, "indirect": None},
    ]

    # Alternating row backgrounds
    row_bgs = ["F8FAFF", WHITE, "F8FAFF"]

    for i, prod in enumerate(product_rows):
        r   = start_row + 2 + i
        bg  = row_bgs[i]
        ws.row_dimensions[r].height = 36

        # A — Ref
        label_cell(ws.cell(r, 1), prod["ref"],
                   bold=True, color=HEADER_DARK, bg="EEF2FF", halign="center")

        # B — CN Code
        input_cell(ws.cell(r, 2), prod["cn_code"] or "", bg_color=MANDATORY_BG, halign="center")

        # C — Product description
        input_cell(ws.cell(r, 3), prod["desc"] or "", bg_color=bg)

        # D — Annual production
        if prod["production"] is not None:
            input_cell(ws.cell(r, 4), prod["production"],
                       bg_color=MANDATORY_BG, halign="right", fmt="#,##0.00")
        else:
            input_cell(ws.cell(r, 4), None, bg_color=MANDATORY_BG, halign="right")

        # E — Direct specific emissions (Scope 1)
        if prod["direct"] is not None:
            calculated_cell(ws.cell(r, 5), prod["direct"],
                            halign="right", fmt="0.0000")
        else:
            input_cell(ws.cell(r, 5), None, bg_color=MANDATORY_BG, halign="right", fmt="0.0000")

        # F — Indirect specific emissions (Scope 2)
        if prod["indirect"] is not None:
            calculated_cell(ws.cell(r, 6), prod["indirect"],
                            halign="right", fmt="0.0000")
        else:
            input_cell(ws.cell(r, 6), None, bg_color=MANDATORY_BG, halign="right", fmt="0.0000")

        # G — Total specific (formula: E+F only if both numeric)
        e_col = f"E{r}"
        f_col = f"F{r}"
        g_formula = f"=IF(AND(ISNUMBER({e_col}),ISNUMBER({f_col})),{e_col}+{f_col},\"\")"
        calculated_cell(ws.cell(r, 7), g_formula, halign="right", fmt="0.0000")

        # H — EU Benchmark (input — looked up from benchmarks_cbam.xlsx by operator)
        input_cell(ws.cell(r, 8), None, bg_color=MANDATORY_BG, halign="right", fmt="0.0000")

        # I — Compliance (formula: compare G vs H, conditional color)
        g_col = f"G{r}"
        h_col = f"H{r}"
        i_formula = (
            f'=IF({g_col}="","",IF({g_col}<={h_col},"✅ COMPLIANT","⚠️ EXCEEDS BENCHMARK"))'
        )
        i_cell = ws.cell(r, 9)
        calculated_cell(i_cell, i_formula, halign="center")
        # Static background — Excel conditional formatting would override at runtime;
        # we pre-fill with neutral gray (openpyxl can't evaluate formulas at write time)
        i_cell.fill = _fill(READONLY_BG)

    # ── Footer note ────────────────────────────────────────────────────────
    note_row = start_row + 5
    note_cell(
        ws, note_row, 1, 9,
        "ℹ  Source: CarbonIQ /cbam/summary · benchmarks_cbam.xlsx · "
        "Each CN code pulls from company_profile + emissions aggregation. "
        "Column G (Total) and I (Compliance) are Excel formulas — auto-calculated when file opens.",
    )

    return note_row + 1


def build_section_3_emissions_data(ws, start_row: int, profile: dict, raw: dict) -> int:
    """
    SECTION 3 — EMISSIONS DATA
    Direct (Scope 1) + Indirect (Scope 2) emissions with specific intensities.

    Args:
        ws         : openpyxl Worksheet
        start_row  : first row to write into
        profile    : dict from _get_profile(conn)
        raw        : dict from _get_emissions_summary(conn, year)

    Returns:
        next_row   : start_row + 6 (title + header + 3 data rows + note)

    Columns:
        A(4)  Ref      B(28) Category      C(18) Annual Total(tCO₂e)
        D(14) Production(t)  E(16) Specific(tCO₂e/t) — formula
        F(10) Scope    G(22) Calc Factor   H(14) Source    I(24) Notes
    """
    HEADER_MID = "2E5496"

    # Column widths
    for col_letter, width in [("A", 4),  ("B", 28), ("C", 18), ("D", 14),
                               ("E", 16), ("F", 10), ("G", 22), ("H", 14), ("I", 24)]:
        ws.column_dimensions[col_letter].width = width

    # ── Section title ──────────────────────────────────────────────────────
    merge_and_style_section_title(
        ws, start_row, 1, 9,
        "SECTION 3 — EMISSIONS DATA",
        bg_color=HEADER_DARK,
    )

    # ── Column headers ─────────────────────────────────────────────────────
    h_row = start_row + 1
    for col, text in enumerate([
        "#", "Emissions Category", "Annual Total\n(tCO₂e)",
        "Production\n(tonnes)", "Specific\n(tCO₂e/t)",
        "Scope", "Calculation Factor", "Source", "Notes",
    ], 1):
        header_cell(ws.cell(h_row, col), text,
                    bg_color=HEADER_MID, fg_color=WHITE,
                    bold=True, size=9, halign="center", wrap=True)
    ws.row_dimensions[h_row].height = 42

    # ── Real values from DB ────────────────────────────────────────────────
    production = float(profile.get("production_annuelle_tonnes") or 0)
    scope1_t   = round(raw.get("scope1_kg", 0) / 1000.0, 4)
    scope2_t   = round(raw.get("scope2_kg", 0) / 1000.0, 4)
    has_data   = raw.get("activity_count", 0) > 0

    # Row numbers — stored for cross-row formula references
    r_scope1 = start_row + 2
    r_scope2 = start_row + 3
    r_total  = start_row + 4

    # ── Row definitions ────────────────────────────────────────────────────
    rows_cfg = [
        {
            "r":        r_scope1,
            "ref":      "3.1",
            "category": "Direct Emissions (Scope 1)",
            "total":    scope1_t if has_data else None,
            "scope":    "Scope 1",
            "factor":   "Fuel EF × consumption + process emissions",
            "source":   "CarbonIQ activities",
            "notes":    "Combustion + calcination",
            "is_total": False,
        },
        {
            "r":        r_scope2,
            "ref":      "3.2",
            "category": "Indirect Emissions (Scope 2)",
            "total":    scope2_t if has_data else None,
            "scope":    "Scope 2",
            "factor":   "0.625 kgCO₂/kWh (Morocco ONEE)",
            "source":   "CarbonIQ activities",
            "notes":    "Grid electricity only",
            "is_total": False,
        },
        {
            "r":        r_total,
            "ref":      "3.3",
            "category": "TOTAL Embedded Emissions (Scope 1 + 2)",
            "total":    None,          # formula
            "scope":    "S1+S2",
            "factor":   "Sum of above",
            "source":   "Calculated",
            "notes":    "Auto-calculated — do not edit",
            "is_total": True,
        },
    ]

    for cfg in rows_cfg:
        r       = cfg["r"]
        is_tot  = cfg["is_total"]
        row_bg  = GREEN_OK if is_tot else WHITE
        ws.row_dimensions[r].height = 36

        # A — Ref
        label_cell(ws.cell(r, 1), cfg["ref"],
                   bold=True, color=HEADER_DARK, bg="EEF2FF", halign="center")

        # B — Category
        b_cell = ws.cell(r, 2)
        b_cell.value     = cfg["category"]
        b_cell.font      = Font(name="Arial", size=10,
                                bold=is_tot, color="000000")
        b_cell.fill      = _fill(GREEN_OK if is_tot else "F8FAFF")
        b_cell.alignment = Alignment(horizontal="left", vertical="center")
        b_cell.border    = thin_border()

        # C — Annual total (tCO₂e): input for S1/S2, formula for TOTAL
        c_cell = ws.cell(r, 3)
        if is_tot:
            calculated_cell(c_cell,
                            f"=C{r_scope1}+C{r_scope2}",
                            halign="right", fmt="#,##0.0000", bg=GREEN_OK)
        elif cfg["total"] is not None:
            calculated_cell(c_cell, cfg["total"],
                            halign="right", fmt="#,##0.0000")
        else:
            input_cell(c_cell, None,
                       bg_color=MANDATORY_BG, halign="right", fmt="#,##0.0000")

        # D — Production (t): input for S1, formula for S2 and TOTAL
        d_cell = ws.cell(r, 4)
        if r == r_scope1:
            if production > 0:
                input_cell(d_cell, production,
                           bg_color=MANDATORY_BG, halign="right", fmt="#,##0.00")
            else:
                input_cell(d_cell, None,
                           bg_color=MANDATORY_BG, halign="right", fmt="#,##0.00")
        else:
            # S2 and TOTAL mirror S1's production cell
            calculated_cell(d_cell, f"=D{r_scope1}",
                            halign="right", fmt="#,##0.00",
                            bg=GREEN_OK if is_tot else READONLY_BG)

        # E — Specific intensity (tCO₂e/t): formula =IF(D=0,"",C/D)
        calculated_cell(ws.cell(r, 5),
                        f'=IF(D{r}=0,"",C{r}/D{r})',
                        halign="right", fmt="0.0000",
                        bg=GREEN_OK if is_tot else READONLY_BG)

        # F — Scope label
        label_cell(ws.cell(r, 6), cfg["scope"],
                   bold=is_tot, color="003399", bg=row_bg, halign="center")

        # G — Calculation factor
        label_cell(ws.cell(r, 7), cfg["factor"],
                   color="595959", bg=row_bg, size=9)

        # H — Source
        label_cell(ws.cell(r, 8), cfg["source"],
                   color="595959", bg=row_bg, size=9, halign="center")

        # I — Notes
        i_cell = ws.cell(r, 9)
        i_cell.value     = cfg["notes"]
        i_cell.font      = Font(name="Arial", size=9, color="595959",
                                italic=is_tot)
        i_cell.fill      = _fill(row_bg)
        i_cell.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True)
        i_cell.border    = thin_border()

    # ── Footer note ────────────────────────────────────────────────────────
    note_row = r_total + 1
    note_cell(
        ws, note_row, 1, 9,
        "ℹ  Source: CarbonIQ /summary · Scope 1 = fuel + process combustion · "
        "Scope 2 = electricity (Morocco ONEE 0.625 kgCO₂/kWh) · "
        "Column E (Specific) and TOTAL row are Excel formulas — auto-calculated.",
    )

    return note_row + 1


def build_section_4_methodology(ws, start_row: int, profile: dict) -> int:
    """
    SECTION 4 — METHODOLOGY
    Documents calculation approach and standards used.

    Args:
        ws         : openpyxl Worksheet
        start_row  : first row to write into
        profile    : dict from _get_profile(conn)

    Returns:
        next_row   : start_row + 7 (title + header + 4 data rows + note)

    Columns:
        A(4)  Ref    B(20) Parameter    C(24) Value
        D(24) Accepted Values           E(20) Source
    """
    HEADER_MID = "2E5496"

    # Column widths
    for col_letter, width in [("A", 4), ("B", 20), ("C", 36),
                               ("D", 28), ("E", 24)]:
        ws.column_dimensions[col_letter].width = width

    # ── Section title ──────────────────────────────────────────────────────
    merge_and_style_section_title(
        ws, start_row, 1, 5,
        "SECTION 4 — METHODOLOGY",
        bg_color=HEADER_DARK,
    )

    # ── Column headers ─────────────────────────────────────────────────────
    h_row = start_row + 1
    for col, text in enumerate(
        ["#", "Parameter", "Value", "Accepted Values", "Source"], 1
    ):
        header_cell(ws.cell(h_row, col), text,
                    bg_color=HEADER_MID, fg_color=WHITE,
                    bold=True, size=9, halign="center", wrap=True)
    ws.row_dimensions[h_row].height = 40

    # ── 4 parameter rows ───────────────────────────────────────────────────
    # Each: (ref, parameter, value_from_db, accepted_values, source_note, is_fixed)
    params = [
        (
            "4.1",
            "Calculation Method",
            profile.get("methode_saisie") or "Calculation-based",
            "Calculation-based / Measurement-based / Mixed",
            "company_profile.methode_saisie",
            False,
        ),
        (
            "4.2",
            "GHG Protocol Version",
            "GHG Protocol Corporate Standard Rev. 2024",
            "GHG Protocol / ISO 14064-1",
            "Fixed standard (2024)",
            True,                      # fixed — not editable
        ),
        (
            "4.3",
            "Monitoring Approach",
            profile.get("monitoring_approach") or "Tier 2 (energy balance)",
            "Tier 1 / Tier 2 / Tier 3",
            "company_profile or default",
            False,
        ),
        (
            "4.4",
            "Verification Status",
            profile.get("statut_verification") or "Self-declared (Draft)",
            "Self-declared / 3rd-party verified",
            "company_profile.statut_verification",
            False,
        ),
    ]

    row_bgs = ["F8FAFF", WHITE, "F8FAFF", WHITE]

    for i, (ref, param, value, accepted, source, is_fixed) in enumerate(params):
        r  = start_row + 2 + i
        bg = row_bgs[i]
        ws.row_dimensions[r].height = 34

        # A — Ref
        label_cell(ws.cell(r, 1), ref,
                   bold=True, color=HEADER_DARK, bg="EEF2FF", halign="center")

        # B — Parameter label
        label_cell(ws.cell(r, 2), param,
                   bold=True, color="000000", bg="F8FAFF")

        # C — Value: fixed (gray, read-only) or input (yellow, blue text)
        c_cell = ws.cell(r, 3)
        if is_fixed:
            calculated_cell(c_cell, value, halign="left")
        else:
            input_cell(c_cell, value, bg_color=MANDATORY_BG)

        # D — Accepted values (white, gray text, size 9)
        d_cell = ws.cell(r, 4)
        d_cell.value     = accepted
        d_cell.font      = Font(name="Arial", size=9, color="737373", italic=True)
        d_cell.fill      = _fill(WHITE)
        d_cell.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True)
        d_cell.border    = thin_border()

        # E — Source (white, darker gray)
        e_cell = ws.cell(r, 5)
        e_cell.value     = source
        e_cell.font      = Font(name="Arial", size=9, color="404040")
        e_cell.fill      = _fill(WHITE)
        e_cell.alignment = Alignment(horizontal="left", vertical="center")
        e_cell.border    = thin_border()

    # ── Footer note ────────────────────────────────────────────────────────
    note_row = start_row + 6
    note_cell(
        ws, note_row, 1, 5,
        "ℹ  No formulas in this section — documents HOW data was calculated, "
        "not the calculations themselves. Fixed values cannot be edited. "
        "Fields marked in yellow can be updated in CarbonIQ → Paramètres.",
    )

    return note_row + 1


def build_section_5_data_quality(ws, start_row: int, profile: dict, raw: dict) -> int:
    """
    SECTION 5 — DATA QUALITY
    Data completeness and measurement coverage indicators.

    Args:
        ws         : openpyxl Worksheet
        start_row  : first row to write into
        profile    : dict from _get_profile(conn)
        raw        : dict from _get_emissions_summary(conn, year)

    Returns:
        next_row   : start_row + 8 (title + header + 5 data rows + note)

    Columns:
        A(4)  Ref    B(26) Indicator    C(12) Value    D(6) Unit
        E(16) Threshold    F(12) Status (formula)
        G(18) Source       H(14) Notes
    """
    HEADER_MID = "2E5496"

    for col_letter, width in [("A", 4),  ("B", 26), ("C", 12), ("D", 6),
                               ("E", 16), ("F", 12), ("G", 18), ("H", 14)]:
        ws.column_dimensions[col_letter].width = width

    # ── Section title ──────────────────────────────────────────────────────
    merge_and_style_section_title(
        ws, start_row, 1, 8,
        "SECTION 5 — DATA QUALITY",
        bg_color=HEADER_DARK,
    )

    # ── Column headers ─────────────────────────────────────────────────────
    h_row = start_row + 1
    for col, text in enumerate(
        ["#", "Indicator", "Value", "Unit", "Threshold", "Status", "Source", "Notes"], 1
    ):
        header_cell(ws.cell(h_row, col), text,
                    bg_color=HEADER_MID, fg_color=WHITE,
                    bold=True, size=9, halign="center", wrap=True)
    ws.row_dimensions[h_row].height = 40

    # Row numbers — needed for cross-row formula references
    r51 = start_row + 2   # 5.1  % Default Scope 1
    r52 = start_row + 3   # 5.2  % Default Scope 2
    r53 = start_row + 4   # 5.3  % Measured  Scope 1  (calculated from 5.1)
    r54 = start_row + 5   # 5.4  Verification status
    r55 = start_row + 6   # 5.5  Overall uncertainty

    # Derive initial % default from DB (0 if real data exists, 100 if not)
    has_data         = raw.get("activity_count", 0) > 0
    pct_default_s1   = 0 if has_data else 100
    pct_default_s2   = 0 if has_data else 100
    verif_status     = profile.get("statut_verification") or "Self-declared (Draft)"

    # ── Row definitions ────────────────────────────────────────────────────
    rows_cfg = [
        {
            "r":          r51,
            "ref":        "5.1",
            "indicator":  "% Default Values — Scope 1",
            "value":      pct_default_s1,
            "unit":       "%",
            "threshold":  "< 20 %",
            "f_formula":  f'=IF(C{r51}="","",IF(VALUE(C{r51})<20,"✅ OK","⚠️ REVIEW"))',
            "source":     "CarbonIQ emissions",
            "notes":      "Default if no measured activity",
            "c_type":     "input",
        },
        {
            "r":          r52,
            "ref":        "5.2",
            "indicator":  "% Default Values — Scope 2",
            "value":      pct_default_s2,
            "unit":       "%",
            "threshold":  "< 20 %",
            "f_formula":  f'=IF(C{r52}="","",IF(VALUE(C{r52})<20,"✅ OK","⚠️ REVIEW"))',
            "source":     "CarbonIQ emissions",
            "notes":      "Default if no electricity data",
            "c_type":     "input",
        },
        {
            "r":          r53,
            "ref":        "5.3",
            "indicator":  "% Measured Data — Scope 1",
            "value":      f'=IF(C{r51}="","",100-VALUE(C{r51}))',
            "unit":       "%",
            "threshold":  "> 80 %",
            "f_formula":  None,        # no auto-status for this row
            "source":     f"Calculated from 5.1",
            "notes":      "= 100 − row 5.1",
            "c_type":     "calc",
        },
        {
            "r":          r54,
            "ref":        "5.4",
            "indicator":  "Data Verification Status",
            "value":      verif_status,
            "unit":       "—",
            "threshold":  "3rd-party verified",
            "f_formula":  None,
            "source":     "company_profile.statut_verification",
            "notes":      "Upgrade before EU submission",
            "c_type":     "input",
        },
        {
            "r":          r55,
            "ref":        "5.5",
            "indicator":  "Overall Uncertainty",
            "value":      "± 5",
            "unit":       "± %",
            "threshold":  "< ± 7.5 %",
            "f_formula":  None,
            "source":     "Internal estimate",
            "notes":      "Update after verification",
            "c_type":     "input",
        },
    ]

    row_bgs = ["F8FAFF", WHITE, "F8FAFF", WHITE, "F8FAFF"]

    for i, cfg in enumerate(rows_cfg):
        r  = cfg["r"]
        bg = row_bgs[i]
        ws.row_dimensions[r].height = 34

        # A — Ref
        label_cell(ws.cell(r, 1), cfg["ref"],
                   bold=True, color=HEADER_DARK, bg="EEF2FF", halign="center")

        # B — Indicator label
        label_cell(ws.cell(r, 2), cfg["indicator"],
                   bold=True, color="000000", bg="F8FAFF")

        # C — Value: input (yellow) or calculated (gray)
        c_cell = ws.cell(r, 3)
        if cfg["c_type"] == "calc":
            calculated_cell(c_cell, cfg["value"], halign="right", fmt="0.00")
        else:
            is_numeric = isinstance(cfg["value"], (int, float))
            input_cell(c_cell, cfg["value"],
                       bg_color=MANDATORY_BG, halign="right",
                       fmt="0.00" if is_numeric else None)

        # D — Unit
        label_cell(ws.cell(r, 4), cfg["unit"],
                   color="595959", bg=bg, halign="center", size=9)

        # E — Threshold
        e_cell = ws.cell(r, 5)
        e_cell.value     = cfg["threshold"]
        e_cell.font      = Font(name="Arial", size=9, color="595959", italic=True)
        e_cell.fill      = _fill(bg)
        e_cell.alignment = Alignment(horizontal="center", vertical="center")
        e_cell.border    = thin_border()

        # F — Status (formula for 5.1 & 5.2, dash otherwise)
        f_cell = ws.cell(r, 6)
        if cfg["f_formula"]:
            calculated_cell(f_cell, cfg["f_formula"], halign="center")
        else:
            label_cell(f_cell, "—", color="AAAAAA", bg=READONLY_BG, halign="center", size=9)

        # G — Source
        g_cell = ws.cell(r, 7)
        g_cell.value     = cfg["source"]
        g_cell.font      = Font(name="Arial", size=9, color="404040")
        g_cell.fill      = _fill(bg)
        g_cell.alignment = Alignment(horizontal="left", vertical="center")
        g_cell.border    = thin_border()

        # H — Notes
        h_cell = ws.cell(r, 8)
        h_cell.value     = cfg["notes"]
        h_cell.font      = Font(name="Arial", size=9, color="737373", italic=True)
        h_cell.fill      = _fill(bg)
        h_cell.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True)
        h_cell.border    = thin_border()

    # ── Footer note ────────────────────────────────────────────────────────
    note_row = r55 + 1
    note_cell(
        ws, note_row, 1, 8,
        "ℹ  Rows 5.1 & 5.2: enter % of default values used (0% = fully measured). "
        "Row 5.3 auto-calculates measured %. "
        "Status column (F) updates automatically via Excel formula.",
    )

    return note_row + 1


def build_section_6_carbon_price_paid(ws, start_row: int) -> int:
    """
    SECTION 6 — CARBON PRICE PAID
    Carbon price exposure and Article 9 deductions (EU ETS).

    Morocco context: NO recognised EU ETS → Article 9 deduction = 0 EUR.
    All values are fixed to 0 except Internal Carbon Pricing (optional).

    Args:
        ws         : openpyxl Worksheet
        start_row  : first row to write into

    Returns:
        next_row   : start_row + 9 (title + warning + header + 5 rows + note)

    Columns:
        A(4)  Ref    B(28) Description    C(14) Amount    D(12) Unit
        E(24) Legal Basis    F(8) Coverage    G(28) Notes
    """
    HEADER_MID  = "2E5496"
    ORANGE_WARN = "FCE4D6"

    for col_letter, width in [("A", 4),  ("B", 28), ("C", 14), ("D", 12),
                               ("E", 24), ("F", 8),  ("G", 28)]:
        ws.column_dimensions[col_letter].width = width

    # ── Section title ──────────────────────────────────────────────────────
    merge_and_style_section_title(
        ws, start_row, 1, 7,
        "SECTION 6 — CARBON PRICE PAID",
        bg_color=HEADER_DARK,
    )

    # ── Warning row: Morocco context ───────────────────────────────────────
    warn_row = start_row + 1
    note_cell(
        ws, warn_row, 1, 7,
        "⚠️  Morocco has NO EU-recognised carbon pricing scheme → "
        "Article 9 deduction = 0 EUR. "
        "Update C6.1 or C6.2 if Morocco adopts a national carbon price or joins EU ETS.",
        bg=ORANGE_WARN, fg_color="7B1200",
    )

    # ── Column headers ─────────────────────────────────────────────────────
    h_row = start_row + 2
    for col, text in enumerate(
        ["#", "Description", "Amount", "Unit", "Legal Basis", "Coverage %", "Notes"], 1
    ):
        header_cell(ws.cell(h_row, col), text,
                    bg_color=HEADER_MID, fg_color=WHITE,
                    bold=True, size=9, halign="center", wrap=True)
    ws.row_dimensions[h_row].height = 40

    # Row numbers for cross-row formulas
    r61 = start_row + 3
    r62 = start_row + 4
    r63 = start_row + 5
    r64 = start_row + 6
    r65 = start_row + 7   # TOTAL

    rows_cfg = [
        {
            "r":       r61,
            "ref":     "6.1",
            "desc":    "Carbon Price Paid (EU ETS)",
            "amount":  0,
            "unit":    "EUR",
            "basis":   "EU ETS Directive 2003/87/EC",
            "pct":     "0%",
            "notes":   "Morocco NOT in EU ETS → deduction = 0 EUR (Art. 9 CBAM)",
            "c_type":  "fixed",        # fixed 0 — not editable
            "is_total": False,
        },
        {
            "r":       r62,
            "ref":     "6.2",
            "desc":    "National Carbon Tax (Morocco)",
            "amount":  0,
            "unit":    "EUR",
            "basis":   "N/A — no national carbon tax",
            "pct":     "0%",
            "notes":   "No national carbon price in Morocco (2026)",
            "c_type":  "fixed",
            "is_total": False,
        },
        {
            "r":       r63,
            "ref":     "6.3",
            "desc":    "Internal Carbon Pricing",
            "amount":  0,
            "unit":    "EUR",
            "basis":   "Internal scheme",
            "pct":     "N/A",
            "notes":   "Optional — for internal cost allocation only",
            "c_type":  "input",        # editable — optional scheme
            "is_total": False,
        },
        {
            "r":       r64,
            "ref":     "6.4",
            "desc":    "EU ETS Reference Price (2025 avg)",
            "amount":  76.50,
            "unit":    "EUR/tCO₂e",
            "basis":   "EU ETS market",
            "pct":     "N/A",
            "notes":   "Reference only — CBAM certificate price may differ slightly",
            "c_type":  "fixed",
            "is_total": False,
        },
        {
            "r":       r65,
            "ref":     "6.5",
            "desc":    "TOTAL Art. 9 Deduction (to deduct from CBAM cost)",
            "amount":  f"=C{r61}+C{r62}",   # formula: ETS + national tax
            "unit":    "EUR",
            "basis":   "Art. 9 CBAM Regulation",
            "pct":     "0%",
            "notes":   "Auto-calculated — Morocco = 0 EUR (no EU ETS)",
            "c_type":  "total",
            "is_total": True,
        },
    ]

    row_bgs = ["F8FAFF", WHITE, "F8FAFF", WHITE]   # alternating for 6.1–6.4

    for i, cfg in enumerate(rows_cfg):
        r       = cfg["r"]
        is_tot  = cfg["is_total"]
        row_bg  = GREEN_OK if is_tot else row_bgs[i % len(row_bgs)]
        ws.row_dimensions[r].height = 34

        # A — Ref
        label_cell(ws.cell(r, 1), cfg["ref"],
                   bold=True, color=HEADER_DARK, bg="EEF2FF", halign="center")

        # B — Description
        b_cell = ws.cell(r, 2)
        b_cell.value     = cfg["desc"]
        b_cell.font      = Font(name="Arial", size=10,
                                bold=is_tot, color="000000")
        b_cell.fill      = _fill(GREEN_OK if is_tot else "F8FAFF")
        b_cell.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True)
        b_cell.border    = thin_border()

        # C — Amount
        c_cell = ws.cell(r, 3)
        if cfg["c_type"] == "total":
            calculated_cell(c_cell, cfg["amount"],
                            halign="right", fmt="#,##0.00", bg=GREEN_OK)
        elif cfg["c_type"] == "fixed":
            calculated_cell(c_cell, cfg["amount"],
                            halign="right", fmt="#,##0.00")
        else:   # input
            input_cell(c_cell, cfg["amount"],
                       bg_color=MANDATORY_BG, halign="right", fmt="#,##0.00")

        # D — Unit
        label_cell(ws.cell(r, 4), cfg["unit"],
                   color="595959", bg=row_bg, halign="center", size=9)

        # E — Legal basis
        e_cell = ws.cell(r, 5)
        e_cell.value     = cfg["basis"]
        e_cell.font      = Font(name="Arial", size=9, color="404040", italic=True)
        e_cell.fill      = _fill(row_bg)
        e_cell.alignment = Alignment(horizontal="left", vertical="center")
        e_cell.border    = thin_border()

        # F — Coverage %
        label_cell(ws.cell(r, 6), cfg["pct"],
                   color="595959", bg=row_bg, halign="center", size=9)

        # G — Notes
        g_cell = ws.cell(r, 7)
        g_cell.value     = cfg["notes"]
        g_cell.font      = Font(name="Arial", size=9, color="595959",
                                italic=is_tot)
        g_cell.fill      = _fill(row_bg)
        g_cell.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True)
        g_cell.border    = thin_border()

    # ── Footer note ────────────────────────────────────────────────────────
    note_row = r65 + 1
    note_cell(
        ws, note_row, 1, 7,
        "ℹ  Art. 9 CBAM Regulation: the CBAM declarant may deduct carbon price "
        "effectively paid in the country of origin. Morocco = 0 EUR. "
        "Row 6.4 (EU ETS price) is for reference only and does NOT affect the deduction formula.",
    )

    return note_row + 1


def build_section_7_sector_specific_params(ws, start_row: int, profile: dict) -> int:
    """
    SECTION 7 — SECTOR-SPECIFIC PARAMETERS
    Documents sector parameters that affect CBAM exposure calculation.

    All 8 rows are written for every company; rows not matching the company's
    sector are greyed out so the file stays reusable as a template.

    Args:
        ws         : openpyxl Worksheet
        start_row  : first row to write into
        profile    : dict from _get_profile(conn)

    Returns:
        next_row   : start_row + 11 (title + header + 8 rows + note)

    Columns:
        A(4)  Ref    B(14) Sector    C(24) Parameter    D(22) Value
        E(8)  Unit   F(14) EU Benchmark    G(18) DB Field    H(18) Required if
    """
    HEADER_MID = "2E5496"

    # Sector background tints
    SECTOR_BG = {
        "cement":     "E8F5E9",
        "steel":      "E3F2FD",
        "aluminium":  "FFF3E0",
        "fertiliser": "F3E5F5",
        "hydrogen":   "E0F7FA",
    }
    INACTIVE_BG = "F5F5F5"   # gray-out rows that don't match company sector

    for col_letter, width in [("A", 4),  ("B", 14), ("C", 24), ("D", 22),
                               ("E", 8),  ("F", 14), ("G", 18), ("H", 18)]:
        ws.column_dimensions[col_letter].width = width

    # ── Section title ──────────────────────────────────────────────────────
    merge_and_style_section_title(
        ws, start_row, 1, 8,
        "SECTION 7 — SECTOR-SPECIFIC PARAMETERS",
        bg_color=HEADER_DARK,
    )

    # ── Column headers ─────────────────────────────────────────────────────
    h_row = start_row + 1
    for col, text in enumerate(
        ["#", "Sector", "Parameter", "Value", "Unit",
         "EU Benchmark", "DB Field", "Required if"], 1
    ):
        header_cell(ws.cell(h_row, col), text,
                    bg_color=HEADER_MID, fg_color=WHITE,
                    bold=True, size=9, halign="center", wrap=True)
    ws.row_dimensions[h_row].height = 40

    # ── Detect active sector ───────────────────────────────────────────────
    raw_sector  = (profile.get("secteur") or "").lower()
    active_key  = None
    if "ciment" in raw_sector or "cement" in raw_sector:
        active_key = "cement"
    elif "acier" in raw_sector or "steel" in raw_sector or "fer" in raw_sector:
        active_key = "steel"
    elif "alumin" in raw_sector or "alu" in raw_sector:
        active_key = "aluminium"
    elif "engrais" in raw_sector or "fertil" in raw_sector:
        active_key = "fertiliser"
    elif "hydrog" in raw_sector:
        active_key = "hydrogen"

    # ── Row definitions ────────────────────────────────────────────────────
    # (ref, sector_key, sector_label, parameter, db_field, unit, benchmark, required_if)
    rows_cfg = [
        ("7.1", "cement",     "Cement",     "Clinker-to-Cement Ratio",
         "clinker_ratio",   "ratio", "0.737",           "Cement exported"),
        ("7.2", "cement",     "Cement",     "EN 197-1 Cement Type",
         "ciment_type",     "—",     "CEM I – CEM V",   "Cement exported"),
        ("7.3", "steel",      "Steel",      "Production Route",
         "route_production","—",     "BF-BOF / EAF",    "Steel exported"),
        ("7.4", "steel",      "Steel",      "Scrap Input Ratio",
         "scrap_ratio",     "%",     "—",               "Steel exported"),
        ("7.5", "aluminium",  "Aluminium",  "Primary vs Secondary",
         "alu_type",        "—",     "Primary / Secondary", "Aluminium exported"),
        ("7.6", "aluminium",  "Aluminium",  "Alloy Content",
         "alloy_pct",       "%",     "< 15 %",          "Aluminium exported"),
        ("7.7", "fertiliser", "Fertiliser", "Ammonia Production Route",
         "engrais_type",    "—",     "SMR / Electrolysis", "Fertiliser exported"),
        ("7.8", "hydrogen",   "Hydrogen",   "H₂ Production Method",
         "h2_methode",      "—",     "SMR / Electrolysis", "Hydrogen exported"),
    ]

    for i, (ref, sector_key, sector_label, param, db_field, unit, benchmark, req_if) in enumerate(rows_cfg):
        r          = start_row + 2 + i
        is_active  = (sector_key == active_key)
        sector_bg  = SECTOR_BG[sector_key]
        row_bg     = sector_bg if is_active else INACTIVE_BG
        text_color = "000000" if is_active else "AAAAAA"
        ws.row_dimensions[r].height = 34

        # A — Ref
        label_cell(ws.cell(r, 1), ref,
                   bold=is_active, color=HEADER_DARK if is_active else "AAAAAA",
                   bg="EEF2FF" if is_active else INACTIVE_BG, halign="center")

        # B — Sector label (tinted background)
        b_cell = ws.cell(r, 2)
        b_cell.value     = sector_label
        b_cell.font      = Font(name="Arial", size=9,
                                bold=is_active, color=text_color)
        b_cell.fill      = _fill(sector_bg if is_active else INACTIVE_BG)
        b_cell.alignment = Alignment(horizontal="center", vertical="center")
        b_cell.border    = thin_border()

        # C — Parameter name
        label_cell(ws.cell(r, 3), param,
                   bold=is_active, color=text_color, bg=row_bg)

        # D — Value: yellow input if active sector, gray placeholder if not
        raw_val = profile.get(db_field)
        needs_add = "← ADD" not in db_field  # fields that exist in DB
        display_val = raw_val if raw_val is not None else (
            "" if is_active else "N/A for this sector"
        )
        d_cell = ws.cell(r, 4)
        if is_active:
            input_cell(d_cell, display_val, bg_color=MANDATORY_BG)
        else:
            calculated_cell(d_cell, display_val, halign="left", bg=INACTIVE_BG)
            d_cell.font = Font(name="Arial", size=9, color="AAAAAA", italic=True)

        # E — Unit
        label_cell(ws.cell(r, 5), unit,
                   color=text_color, bg=row_bg, halign="center", size=9)

        # F — EU Benchmark
        f_cell = ws.cell(r, 6)
        f_cell.value     = benchmark
        f_cell.font      = Font(name="Arial", size=9,
                                color="003399" if is_active else "BBBBBB",
                                bold=is_active)
        f_cell.fill      = _fill(row_bg)
        f_cell.alignment = Alignment(horizontal="center", vertical="center")
        f_cell.border    = thin_border()

        # G — DB field
        needs_add_flag = db_field in (
            "clinker_ratio", "scrap_ratio", "alloy_pct"
        )
        g_text  = f"{db_field}" + (" ← ADD" if needs_add_flag else "")
        g_color = "CC0000" if (needs_add_flag and is_active) else ("595959" if is_active else "CCCCCC")
        g_cell  = ws.cell(r, 7)
        g_cell.value     = g_text
        g_cell.font      = Font(name="Arial", size=9, color=g_color,
                                italic=needs_add_flag)
        g_cell.fill      = _fill(row_bg)
        g_cell.alignment = Alignment(horizontal="left", vertical="center")
        g_cell.border    = thin_border()

        # H — Required if
        label_cell(ws.cell(r, 8), req_if,
                   color=text_color, bg=row_bg, size=9,
                   bold=is_active)

    # ── Footer note ────────────────────────────────────────────────────────
    note_row = start_row + 10
    note_cell(
        ws, note_row, 1, 8,
        "ℹ  Fill ONLY rows relevant to your exported sector — leave others blank. "
        "Active sector rows are highlighted. "
        "Fields marked '← ADD' require adding the column to company_profile table in CarbonIQ.",
    )

    return note_row + 1


# ── Workbook Builder ──────────────────────────────────────────────────────────

def generate_communication_template_excel(year: int = 2026, user_id: int = 0) -> Workbook:
    """
    Builds the complete Summary_Communication Excel workbook.

    Opens its own DB connection, calls all 7 section builders in sequence,
    adds global header (rows 1–5) and footer, configures print/view settings.

    Returns:
        Workbook — ready to save or stream. Caller is responsible for closing.

    Raises:
        HTTPException 404 if company_profile not found.
        HTTPException 500 on any other error.
    """
    conn = get_connection()
    try:
        profile = _get_profile(conn, user_id)    # raises 404 if not found
        raw     = _get_emissions_summary(conn, year)
    finally:
        conn.close()

    wb = Workbook()
    wb.remove(wb.active)

    # ── ⚠️  TAB NAME MANDATORY — do NOT rename ────────────────────────────────
    ws = wb.create_sheet("Summary_Communication")

    # ── Sheet view ────────────────────────────────────────────────────────────
    ws.sheet_view.showGridLines = False

    # ── Print settings ────────────────────────────────────────────────────────
    ws.print_title_rows              = "1:7"
    ws.page_setup.orientation        = "landscape"
    ws.page_setup.fitToPage          = True
    ws.page_setup.fitToWidth         = 1
    ws.page_margins.left             = 0.5
    ws.page_margins.right            = 0.5
    ws.page_margins.top              = 0.5
    ws.page_margins.bottom           = 0.5

    # ── Rows 1–5: Global EU banner ────────────────────────────────────────────
    #   Row 1 — EU flag banner
    ws.merge_cells("A1:I1")
    _styled(ws["A1"],
            f"🇪🇺  EU CBAM — Official Communication Template  ·  Reporting Year {year}",
            fill_hex=EU_BLUE, bold=True, color=EU_YELLOW, size=14,
            h_align="center", border=False)
    ws.row_dimensions[1].height = 30

    #   Row 2 — Regulation reference
    ws.merge_cells("A2:I2")
    _styled(ws["A2"],
            "Regulation (EU) 2023/956  ·  Implementing Regulation (EU) 2025/2620  "
            "·  IR 2025/2547 (Verification)  ·  Annex I — CN Codes",
            fill_hex=HEADER_DARK, bold=False, color=WHITE, size=10,
            h_align="center", italic=True, border=False)
    ws.row_dimensions[2].height = 18

    #   Row 3 — Tab name warning
    ws.merge_cells("A3:I3")
    _styled(ws["A3"],
            '⚠️  MANDATORY: This tab MUST be named  "Summary_Communication"  — '
            "renaming causes automatic rejection by the EU CBAM Portal.",
            fill_hex=EU_YELLOW, bold=True, color="CC0000", size=10,
            h_align="center", wrap=True, border=False)
    ws.row_dimensions[3].height = 22

    #   Row 4 — Operator + date meta-line
    ws.merge_cells("A4:I4")
    _styled(ws["A4"],
            f"Operator: {profile.get('nom_entreprise', '[operator name]')}  "
            f"·  Sector: {profile.get('secteur', '')}  "
            f"·  CN Code: {profile.get('cn_code', '')}  "
            f"·  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            fill_hex="EFF3FB", bold=False, color="1F3864", size=9,
            h_align="left", border=False)
    ws.row_dimensions[4].height = 16

    #   Row 5 — Blank separator
    ws.row_dimensions[5].height = 8

    # ── Section builders (chained via next_row) ───────────────────────────────
    next_row = 6
    next_row = build_section_1_installation_details(ws, next_row, profile)
    next_row = build_section_2_summary_of_products(ws, next_row, profile, raw)
    next_row = build_section_3_emissions_data(ws, next_row, profile, raw)
    next_row = build_section_4_methodology(ws, next_row, profile)
    next_row = build_section_5_data_quality(ws, next_row, profile, raw)
    next_row = build_section_6_carbon_price_paid(ws, next_row)
    next_row = build_section_7_sector_specific_params(ws, next_row, profile)

    # ── Footer ────────────────────────────────────────────────────────────────
    footer_row = next_row + 1
    ws.merge_cells(f"A{footer_row}:I{footer_row}")
    _styled(ws[f"A{footer_row}"],
            "Generated by CarbonIQ MRV Platform  ·  "
            "Conforme EU 2023/956 + 2025/2620 + IR 2025/2547  ·  "
            "CBAM Portal: https://customs.ec.europa.eu/cbam",
            fill_hex=EU_BLUE, bold=False, color=WHITE, size=8,
            h_align="center", italic=True, border=False)
    ws.row_dimensions[footer_row].height = 18

    return wb


# ── FastAPI Endpoint ───────────────────────────────────────────────────────────

@router.get(
    "/communication-template-v2",
    summary="EU CBAM Communication Template Excel — Full 7-Section (UE 2023/956 + 2025/2620)",
    response_description="Fichier .xlsx conforme EU CBAM Portal — 7 sections complètes",
)
def download_communication_template(
    year: int = Query(2026, ge=2026, le=2030, description="Année de reporting CBAM"),
    current_user: dict = Depends(get_current_user),
):
    """
    Génère le Communication Template Excel CBAM complet avec 7 sections.

    ⚠️ L'onglet **Summary_Communication** ne doit JAMAIS être renommé.

    Sections incluses :
    1. Installation Details (identification opérateur)
    2. Summary of Products (CN codes + émissions spécifiques + benchmarks)
    3. Emissions Data (Scope 1, Scope 2, TOTAL avec formules)
    4. Methodology (méthode de calcul, GHG Protocol, vérification)
    5. Data Quality (% valeurs par défaut, incertitude)
    6. Carbon Price Paid (Article 9 — Maroc = 0 EUR)
    7. Sector-Specific Parameters (Ciment/Acier/Aluminium/Engrais/Hydrogène)
    """
    try:
        wb  = generate_communication_template_excel(year, current_user["user_id"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"CBAM_Communication_Template_{year}.xlsx"
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
