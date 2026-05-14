"""
routers/cbam_communication.py

Endpoint FastAPI pour générer le Communication Template Excel CBAM officiel
avec les données réelles de l'installation marocaine.
"""

from fastapi import APIRouter, HTTPException, Query, Depends
from fastapi.responses import StreamingResponse, JSONResponse
from dependencies import get_current_user
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import get_connection

router = APIRouter(prefix="/cbam", tags=["CBAM Communication"])

# ── Styles Excel ───────────────────────────────────────────────────────────────
def get_styles():
    return {
        "header_font":    Font(bold=True, color="FFFFFF", size=11),
        "header_fill":    PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
        "subheader_fill": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        "subheader_font": Font(bold=True, size=10),
        "border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin")
        ),
        "center_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left_align":   Alignment(horizontal="left",   vertical="center"),
        "right_align":  Alignment(horizontal="right",  vertical="center"),
    }

# ── CN code → secteur / groupe ─────────────────────────────────────────────────
CN_CODE_MAPPING = {
    # ACIER / FER
    "72011000": {"secteur": "Steel", "groupe": "A", "description": "Iron & Steel - Fonte brute"},
    "72011010": {"secteur": "Steel", "groupe": "A", "description": "Iron & Steel - Flat products"},
    "72011011": {"secteur": "Steel", "groupe": "A", "description": "Iron & Steel - Long products"},
    # ALUMINIUM
    "76011000": {"secteur": "Aluminium", "groupe": "A", "description": "Primary aluminium"},
    "76013000": {"secteur": "Aluminium", "groupe": "A", "description": "Aluminium alloys"},
    # CIMENT
    "25232100": {"secteur": "Cement",    "groupe": "B", "description": "Portland cement"},
    "25232900": {"secteur": "Cement",    "groupe": "B", "description": "Other cements"},
    "25239000": {"secteur": "Cement",    "groupe": "B", "description": "Other hydraulic cements"},
    # ENGRAIS
    "28141000": {"secteur": "Fertiliser","groupe": "B", "description": "Anhydrous ammonia"},
    "31021000": {"secteur": "Fertiliser","groupe": "B", "description": "Urea"},
    "31023000": {"secteur": "Fertiliser","groupe": "B", "description": "Ammonium nitrate"},
    "31024000": {"secteur": "Fertiliser","groupe": "B", "description": "Ammonium sulphate"},
    # HYDROGÈNE
    "28042000": {"secteur": "Hydrogen",  "groupe": "A", "description": "Hydrogen"},
}

# ── Helpers DB ─────────────────────────────────────────────────────────────────

def _get_profile(conn, user_id: int) -> dict:
    cursor = conn.cursor()
    cursor.execute(
        "SELECT * FROM company_profile WHERE user_id = %s ORDER BY id LIMIT 1",
        (user_id,)
    )
    row = cursor.fetchone()
    cursor.close()
    if not row:
        raise HTTPException(status_code=404, detail="Profil entreprise non configuré — allez dans Paramètres")
    return dict(row)


def _get_emissions_summary(conn, year: int) -> dict:
    """Somme Scope 1 et Scope 2 pour l'année donnée (une seule requête SQL)."""
    cursor = conn.cursor()
    cursor.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN e.scope = 1 THEN e.co2_kg ELSE 0 END), 0) AS scope1_kg,
            COALESCE(SUM(CASE WHEN e.scope = 2 THEN e.co2_kg ELSE 0 END), 0) AS scope2_kg,
            COUNT(DISTINCT a.id) AS activity_count
        FROM emissions e
        JOIN activities a ON e.activity_id = a.id
        WHERE a.actif = true
          AND e.actif = true
          AND EXTRACT(YEAR FROM a.date) = %s
    """, (year,))
    row = cursor.fetchone()
    cursor.close()
    if not row:
        return {"scope1_kg": 0.0, "scope2_kg": 0.0, "activity_count": 0}
    return {
        "scope1_kg":      float(row["scope1_kg"]),
        "scope2_kg":      float(row["scope2_kg"]),
        "activity_count": int(row["activity_count"]),
    }

# ── Utilitaires Excel ──────────────────────────────────────────────────────────

def _header_row(ws, row: int, headers: list, styles: dict):
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row, col, text)
        cell.font      = styles["header_font"]
        cell.fill      = styles["header_fill"]
        cell.alignment = styles["center_align"]
        cell.border    = styles["border"]

def _data_cell(ws, row: int, col: int, value, styles: dict, align: str = "left"):
    cell = ws.cell(row, col, value)
    cell.alignment = styles[f"{align}_align"]
    cell.border    = styles["border"]
    return cell

def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Endpoint principal ─────────────────────────────────────────────────────────

@router.post("/communication-template")
def generate_communication_template(
    year: int = 2026,
    current_user: dict = Depends(get_current_user),
):
    """
    Génère le Communication Template Excel CBAM officiel
    avec les données réelles de l'installation.

    - **year** : Année de reporting (défaut 2026)
    """
    conn = get_connection()
    try:
        profile = _get_profile(conn, current_user["user_id"])
        raw     = _get_emissions_summary(conn, year)
        styles  = get_styles()

        # Calculs dérivés
        production    = float(profile.get("production_annuelle_tonnes") or 1)
        scope1_t      = raw["scope1_kg"] / 1000.0
        scope2_t      = raw["scope2_kg"] / 1000.0
        scope1_spec   = scope1_t / production if production > 0 else 0.0
        scope2_spec   = scope2_t / production if production > 0 else 0.0
        total_spec    = scope1_spec + scope2_spec
        has_data      = raw["activity_count"] > 0

        cn_code = profile.get("cn_code") or ""
        cn_info = CN_CODE_MAPPING.get(cn_code, {
            "secteur":     profile.get("secteur", ""),
            "groupe":      "A/B",
            "description": "Produit",
        })

        wb = Workbook()
        wb.remove(wb.active)

        # ══════════════════════════════════════════════════════════════
        # Feuille 1 — Summary_Communication
        # ══════════════════════════════════════════════════════════════
        ws1 = wb.create_sheet("Summary_Communication", 0)
        _header_row(ws1, 1, [
            "CN Code", "Product Description", "Route/Method",
            "Direct Emissions (tCO₂e/t)", "Indirect Emissions (tCO₂e/t)",
            "Total Emissions (tCO₂e/t)", "Calculation Method",
            "% Default Values Used", "Carbon Price Paid (€/tCO₂e)", "Notes"
        ], styles)

        row_data = [
            cn_code,
            cn_info["description"],
            profile.get("route_production") or "TBD",
            round(scope1_spec, 4),
            round(scope2_spec, 4),
            round(total_spec, 4),
            "Measured" if has_data else "Default values",
            "0%" if has_data else "100%",
            "75.36",
            f"Production : {production:.0f} t/an · Activités : {raw['activity_count']}",
        ]
        for col, val in enumerate(row_data, 1):
            align = "right" if col in (4, 5, 6, 9) else "left"
            _data_cell(ws1, 2, col, val, styles, align)

        _set_col_widths(ws1, [14, 28, 14, 20, 20, 20, 18, 18, 18, 35])

        # ══════════════════════════════════════════════════════════════
        # Feuille 2 — Summary_Products
        # ══════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("Summary_Products", 1)
        _header_row(ws2, 1, [
            "CN Code", "Product Name (FR)", "Product Name (EN)", "Composition",
            "Annual Production (tonnes)", "Route Classification", "HS Chapter"
        ], styles)

        _data_cell(ws2, 2, 1, cn_code,                             styles)
        _data_cell(ws2, 2, 2, cn_info["description"],              styles)
        _data_cell(ws2, 2, 3, cn_info["description"],              styles)
        _data_cell(ws2, 2, 4, "TBD — composition détaillée",       styles)
        _data_cell(ws2, 2, 5, production,                          styles, "right")
        _data_cell(ws2, 2, 6, cn_info["groupe"],                   styles, "center")
        _data_cell(ws2, 2, 7, cn_code[:2] if len(cn_code) >= 2 else "", styles, "center")
        _set_col_widths(ws2, [14, 28, 28, 30, 20, 16, 12])

        # ══════════════════════════════════════════════════════════════
        # Feuille 3 — Summary_Processes
        # ══════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet("Summary_Processes", 2)
        _header_row(ws3, 1, [
            "Process Code", "Process Description", "CN Code",
            "Activity Level (t/year)", "Direct Emissions (tCO₂e/year)",
            "Indirect Emissions (tCO₂e/year)", "Verification Status"
        ], styles)

        _data_cell(ws3, 2, 1, "P001",                                     styles)
        _data_cell(ws3, 2, 2, f"{profile.get('secteur','')} — Primary process", styles)
        _data_cell(ws3, 2, 3, cn_code,                                    styles)
        _data_cell(ws3, 2, 4, production,                                 styles, "right")
        _data_cell(ws3, 2, 5, round(scope1_t, 2),                        styles, "right")
        _data_cell(ws3, 2, 6, round(scope2_t, 2),                        styles, "right")
        _data_cell(ws3, 2, 7, "Measured" if has_data else "Default",     styles, "center")
        _set_col_widths(ws3, [14, 32, 14, 20, 22, 22, 18])

        # ══════════════════════════════════════════════════════════════
        # Feuille 4 — Onglet sectoriel (A_Steel / A_Aluminium / B_Cement…)
        # ══════════════════════════════════════════════════════════════
        secteur = (profile.get("secteur") or "").lower()

        if "acier" in secteur or "steel" in secteur or "fer" in secteur:
            sheet_name = "A_Steel"
            headers4   = [
                "Installation ID", "Steel Product Type", "Production Route",
                "Annual Output (t)", "Direct CO₂ (tCO₂e/year)", "Indirect CO₂ (tCO₂e/year)",
                "Specific Emissions (tCO₂e/t)", "Benchmark (tCO₂e/t)", "Compliance"
            ]
            row4 = ["INST-001", cn_info["description"], profile.get("route_production") or "TBD",
                    production, round(scope1_t,2), round(scope2_t,2),
                    round(total_spec,4), "TBD", "—"]

        elif "aluminium" in secteur or "alu" in secteur:
            sheet_name = "A_Aluminium"
            headers4   = [
                "Installation ID", "Alu Product", "Process Type",
                "Annual Output (t)", "Direct CO₂ (tCO₂e/year)",
                "Specific Emissions (tCO₂e/t)", "PFC Emissions (tCO₂e/t)",
                "Benchmark (tCO₂e/t)", "Compliance"
            ]
            row4 = ["INST-001", cn_info["description"], profile.get("route_production") or "TBD",
                    production, round(scope1_t,2),
                    round(scope1_spec,4), "1.342 (primaire) / 0 (secondaire)",
                    "TBD", "—"]

        elif "ciment" in secteur or "cement" in secteur:
            sheet_name = "B_Cement"
            headers4   = [
                "Installation ID", "Cement Type (EN 197-1)", "Annual Output (t)",
                "Calcination CO₂ (tCO₂e/year)", "Fuel CO₂ (tCO₂e/year)",
                "Electricity CO₂ (tCO₂e/year)",
                "Specific Direct (tCO₂e/t)", "Specific Indirect (tCO₂e/t)",
                "Total Specific (tCO₂e/t)", "Benchmark (tCO₂e/t)", "Compliance"
            ]
            row4 = ["INST-001", profile.get("route_production") or "TBD", production,
                    "TBD", round(scope1_t,2), round(scope2_t,2),
                    round(scope1_spec,4), round(scope2_spec,4), round(total_spec,4),
                    "0.666", "—"]

        elif "engrais" in secteur or "fertiliser" in secteur or "fertilizer" in secteur:
            sheet_name = "B_Fertiliser"
            headers4   = [
                "Installation ID", "Fertiliser Type", "Annual Output (t)",
                "Primary Process CO₂ (tCO₂e/year)", "Secondary Process CO₂ (tCO₂e/year)",
                "Electricity CO₂ (tCO₂e/year)",
                "Specific Direct (tCO₂e/t)", "Specific Indirect (tCO₂e/t)",
                "Total Specific (tCO₂e/t)", "Benchmark (tCO₂e/t)", "Compliance"
            ]
            row4 = ["INST-001", cn_info["description"], production,
                    round(scope1_t,2), "—", round(scope2_t,2),
                    round(scope1_spec,4), round(scope2_spec,4), round(total_spec,4),
                    "TBD", "—"]

        elif "hydrog" in secteur:
            sheet_name = "A_Hydrogen"
            headers4   = [
                "Installation ID", "H₂ Production Method", "Annual Output (t)",
                "Direct CO₂ (tCO₂e/year)", "Specific Emissions (tCO₂e/t)",
                "Benchmark (tCO₂e/t)", "Compliance"
            ]
            row4 = ["INST-001", profile.get("route_production") or "TBD", production,
                    round(scope1_t,2), round(scope1_spec,4), "5.089", "—"]

        else:
            sheet_name = "A_Product"
            headers4   = [
                "Installation ID", "Product Type", "Annual Output (t)",
                "Direct CO₂ (tCO₂e/year)", "Indirect CO₂ (tCO₂e/year)",
                "Specific Emissions (tCO₂e/t)", "Benchmark (tCO₂e/t)", "Compliance"
            ]
            row4 = ["INST-001", cn_info["description"], production,
                    round(scope1_t,2), round(scope2_t,2), round(total_spec,4), "TBD", "—"]

        ws4 = wb.create_sheet(sheet_name, 3)
        _header_row(ws4, 1, headers4, styles)
        for col, val in enumerate(row4, 1):
            align = "right" if isinstance(val, (int, float)) else "left"
            _data_cell(ws4, 2, col, val, styles, align)
        _set_col_widths(ws4, [14] * len(headers4))

        # ══════════════════════════════════════════════════════════════
        # Feuille 5 — E_PurchPrec (précurseurs achetés)
        # ══════════════════════════════════════════════════════════════
        ws5 = wb.create_sheet("E_PurchPrec", 4)
        _header_row(ws5, 1, [
            "Precursor Code", "Precursor Name", "Supplier Country",
            "Annual Quantity (t)", "Embedded CO₂ (tCO₂e/t)",
            "Total Embedded CO₂ (tCO₂e)", "Documentation", "Verification Status"
        ], styles)
        _data_cell(ws5, 2, 1, "[Ajouter les précurseurs achetés]", styles)
        _set_col_widths(ws5, [16, 28, 18, 18, 18, 20, 20, 18])

        # ══════════════════════════════════════════════════════════════
        # Feuille 6 — Metadata
        # ══════════════════════════════════════════════════════════════
        ws6 = wb.create_sheet("Metadata", 5)
        ws6.merge_cells("A1:B1")
        title = ws6["A1"]
        title.value     = "CBAM Communication Template — Metadata"
        title.font      = Font(bold=True, size=12, color="FFFFFF")
        title.fill      = styles["header_fill"]
        title.alignment = styles["center_align"]
        ws6.row_dimensions[1].height = 26

        meta = [
            ("Reporting Entity",          profile.get("nom_entreprise", "")),
            ("Country",                   "Morocco"),
            ("Reporting Year",            year),
            ("Report Date",               datetime.now().strftime("%Y-%m-%d")),
            ("EU Regulation",             "UE 2025/2620 & UE 2025/2621"),
            ("Verification Status",       "Draft"),
            ("CN Code",                   cn_code),
            ("Secteur CBAM",              profile.get("secteur", "")),
            ("Route de production",       profile.get("route_production", "")),
            ("Production annuelle (t/an)",production),
            ("Scope 1 total (tCO₂e)",    round(scope1_t, 2)),
            ("Scope 2 total (tCO₂e)",    round(scope2_t, 2)),
            ("Intensité réelle (tCO₂e/t)",round(total_spec, 4)),
            ("Nombre d'activités",        raw["activity_count"]),
        ]
        for r, (label, value) in enumerate(meta, 2):
            lbl = ws6.cell(r, 1, label)
            lbl.font = Font(bold=True)
            lbl.fill = styles["subheader_fill"]
            lbl.border = styles["border"]
            lbl.alignment = styles["left_align"]
            val = ws6.cell(r, 2, str(value) if value is not None else "")
            val.border    = styles["border"]
            val.alignment = styles["left_align"]
        ws6.column_dimensions["A"].width = 26
        ws6.column_dimensions["B"].width = 42

        # ── Sauvegarde et réponse ──────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        nom = profile.get("nom_entreprise", "Installation").replace(" ", "_")
        filename = f"CBAM_Communication_Template_{nom}_{year}.xlsx"

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        conn.close()


# ── Endpoint prévisualisation JSON ─────────────────────────────────────────────

@router.get("/communication-template-preview")
def preview_communication_template(
    year: int = Query(2026, ge=2026, le=2030),
    current_user: dict = Depends(get_current_user),
):
    """Retourne les données du Communication Template en JSON pour prévisualisation."""
    conn = get_connection()
    try:
        profile = _get_profile(conn, current_user["user_id"])
        raw     = _get_emissions_summary(conn, year)

        production  = float(profile.get("production_annuelle_tonnes") or 1)
        scope1_t    = raw["scope1_kg"] / 1000.0
        scope2_t    = raw["scope2_kg"] / 1000.0
        scope1_spec = round(scope1_t / production, 4) if production > 0 else 0.0
        scope2_spec = round(scope2_t / production, 4) if production > 0 else 0.0
        total_spec  = round(scope1_spec + scope2_spec, 4)
        has_data    = raw["activity_count"] > 0

        cn_code = profile.get("cn_code") or ""
        cn_info = CN_CODE_MAPPING.get(cn_code, {
            "secteur":     profile.get("secteur", ""),
            "groupe":      "A/B",
            "description": "Produit",
        })

        return JSONResponse(content={
            "reporting_year":  year,
            "generation_date": datetime.now().isoformat(),
            "sheets": {
                "Summary_Communication": {
                    "headers": ["CN Code", "Description", "Route", "Direct (tCO₂e/t)",
                                "Indirect (tCO₂e/t)", "Total (tCO₂e/t)", "Method", "% Default"],
                    "rows": [[
                        cn_code, cn_info["description"],
                        profile.get("route_production") or "TBD",
                        scope1_spec, scope2_spec, total_spec,
                        "Measured" if has_data else "Default values",
                        "0%" if has_data else "100%",
                    ]]
                },
                "Summary_Products": {
                    "headers": ["CN Code", "Product Name", "Annual Production (t)", "Route Group"],
                    "rows": [[cn_code, cn_info["description"], production, cn_info["groupe"]]]
                },
                "Summary_Processes": {
                    "headers": ["Process", "Activity Level (t/yr)", "Direct (tCO₂e/yr)",
                                "Indirect (tCO₂e/yr)", "Verification"],
                    "rows": [[
                        f"{profile.get('secteur','')} — Primary process",
                        production, round(scope1_t, 2), round(scope2_t, 2),
                        "Measured" if has_data else "Default",
                    ]]
                },
                "Metadata": {
                    "headers": ["Field", "Value"],
                    "rows": [
                        ["Reporting Entity",     profile.get("nom_entreprise", "")],
                        ["Country",              "Morocco"],
                        ["Reporting Year",       year],
                        ["CN Code",              cn_code],
                        ["Secteur CBAM",         profile.get("secteur", "")],
                        ["Route de production",  profile.get("route_production", "")],
                        ["Production (t/an)",    production],
                        ["Scope 1 (tCO₂e)",      round(scope1_t, 2)],
                        ["Scope 2 (tCO₂e)",      round(scope2_t, 2)],
                        ["Intensité (tCO₂e/t)",  total_spec],
                        ["Nb activités",         raw["activity_count"]],
                        ["EU Regulation",        "UE 2025/2620 & UE 2025/2621"],
                        ["Verification Status",  "Draft"],
                    ]
                },
            }
        })
    finally:
        conn.close()
